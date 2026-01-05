#!/usr/bin/env python3
"""
업체분류 매칭 스크립트
- 9월~12월 파일에서 거래처 → 업체분류 매핑 생성
- 1월~8월 파일의 공란인 업체분류를 매핑으로 채움
"""

from openpyxl import load_workbook
from pathlib import Path
from collections import defaultdict
import shutil
from datetime import datetime

DATA_DIR = Path('data/2025')
BACKUP_DIR = Path('data/2025_backup')

def main():
    print("=" * 60)
    print("업체분류 매칭 스크립트")
    print("=" * 60)

    # 1. 백업 생성
    if not BACKUP_DIR.exists():
        BACKUP_DIR.mkdir(parents=True)
        print(f"[백업] {BACKUP_DIR} 폴더 생성")

    # 2. 9월~12월 파일에서 거래처 → 업체분류 매핑 생성
    print("\n[1단계] 9월~12월 파일에서 업체분류 매핑 수집...")
    company_mapping = {}

    for month in ['09', '10', '11', '12']:
        filename = f'2025_{month}.xlsx'
        filepath = DATA_DIR / filename
        if not filepath.exists():
            print(f"  - {filename}: 파일 없음")
            continue

        wb = load_workbook(filepath, read_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]

        if '거래처' not in headers or '업체분류' not in headers:
            print(f"  - {filename}: 필요한 컬럼 없음")
            wb.close()
            continue

        거래처_idx = headers.index('거래처')
        업체분류_idx = headers.index('업체분류')

        count = 0
        for row in ws.iter_rows(min_row=2):
            거래처 = row[거래처_idx].value
            업체분류 = row[업체분류_idx].value

            if 거래처 and 업체분류:
                거래처 = str(거래처).strip()
                업체분류 = str(업체분류).strip()
                if 거래처 and 업체분류:
                    company_mapping[거래처] = 업체분류
                    count += 1

        wb.close()
        print(f"  - {filename}: {count}개 매핑 수집")

    print(f"\n[매핑 완료] 총 {len(company_mapping)}개 거래처 매핑")

    # 3. 1월~8월 파일 업데이트
    print("\n[2단계] 1월~8월 파일 업체분류 업데이트...")

    total_updated = 0
    for month in ['01', '02', '03', '04', '05', '06', '07', '08']:
        filename = f'2025_{month}.xlsx'
        filepath = DATA_DIR / filename
        if not filepath.exists():
            continue

        # 백업
        backup_path = BACKUP_DIR / f"{filename}.{datetime.now().strftime('%Y%m%d_%H%M%S')}.bak"
        shutil.copy(filepath, backup_path)

        # 읽기 전용이 아닌 모드로 열기
        wb = load_workbook(filepath)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]

        if '거래처' not in headers or '업체분류' not in headers:
            print(f"  - {filename}: 필요한 컬럼 없음")
            wb.close()
            continue

        거래처_idx = headers.index('거래처')
        업체분류_idx = headers.index('업체분류')

        updated = 0
        not_found = 0
        for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
            거래처_cell = row[거래처_idx]
            업체분류_cell = row[업체분류_idx]

            거래처 = str(거래처_cell.value).strip() if 거래처_cell.value else ''
            현재_업체분류 = 업체분류_cell.value

            # 업체분류가 비어있고 매핑에 있으면 업데이트
            if not 현재_업체분류 and 거래처 and 거래처 in company_mapping:
                ws.cell(row=row_num, column=업체분류_idx + 1, value=company_mapping[거래처])
                updated += 1
            elif not 현재_업체분류 and 거래처:
                not_found += 1

        if updated > 0:
            wb.save(filepath)
            total_updated += updated

        wb.close()
        print(f"  - {filename}: {updated}건 업데이트, {not_found}건 매핑 없음")

    print("\n" + "=" * 60)
    print(f"[완료] 총 {total_updated}건 업체분류 업데이트됨")
    print(f"[백업] {BACKUP_DIR} 폴더에 원본 파일 백업됨")
    print("=" * 60)
    print("\n다음 단계: SQLite DB를 업데이트하려면 서버를 재시작하세요")
    print("  python flask_dashboard.py")

if __name__ == '__main__':
    main()
